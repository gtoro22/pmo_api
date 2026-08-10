"""Pruebas del adaptador de exportacion a Excel."""

from __future__ import annotations

from openpyxl import load_workbook

from tracking_goals.domain.model.registro_plano import RegistroPlano
from tracking_goals.domain.services.aplanador_objetivos import AplanadorObjetivos
from tracking_goals.infrastructure.exportacion.exportador_excel import ExportadorExcel
from tracking_goals.infrastructure.http.mapeadores import MapeadorRespuesta


def test_genera_excel_con_encabezado_y_datos(tmp_path, respuesta_documentacion):
    resultado = MapeadorRespuesta().a_resultado(respuesta_documentacion)
    registros = AplanadorObjetivos().aplanar(resultado)

    destino = tmp_path / "sub" / "reporte.xlsx"
    archivo = ExportadorExcel().exportar(registros, destino)

    assert archivo.exists()
    hoja = load_workbook(archivo).active
    assert [celda.value for celda in hoja[1]] == list(RegistroPlano.columnas())
    assert hoja.max_row == 2

    fila = {celda.column_letter: celda.value for celda in hoja[2]}
    assert "5555553333" in fila.values()
    assert "VENEZUELA" in fila.values()


def test_exporta_sin_registros(tmp_path):
    archivo = ExportadorExcel().exportar([], tmp_path / "vacio.xlsx")
    hoja = load_workbook(archivo).active
    assert hoja.max_row == 1


def test_sanea_caracteres_de_control(tmp_path):
    registro = RegistroPlano(objetivo="Texto\x07con control\nsalto")
    archivo = ExportadorExcel().exportar([registro], tmp_path / "control.xlsx")
    hoja = load_workbook(archivo).active
    columna = RegistroPlano.columnas().index("objetivo") + 1
    assert hoja.cell(row=2, column=columna).value == "Textocon control\nsalto"
