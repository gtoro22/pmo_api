"""Adaptador de exportacion a Excel (.xlsx) usando openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tracking_goals.application.ports.exportador_registros import ExportadorRegistros
from tracking_goals.domain.exceptions import ErrorDeExportacion
from tracking_goals.domain.model.registro_plano import RegistroPlano

logger = logging.getLogger(__name__)

NOMBRE_HOJA = "Objetivos"
ANCHO_MINIMO = 12
ANCHO_MAXIMO = 60
CARACTERES_ILEGALES = {chr(codigo) for codigo in range(32)} - {"\n", "\t"}


class ExportadorExcel(ExportadorRegistros):
    """Escribe los registros planos en una unica hoja de calculo."""

    def __init__(self, nombre_hoja: str = NOMBRE_HOJA) -> None:
        self._nombre_hoja = nombre_hoja

    def exportar(self, registros: Sequence[RegistroPlano], destino: Path) -> Path:
        try:
            destino.parent.mkdir(parents=True, exist_ok=True)
            libro = Workbook()
            hoja = libro.active
            hoja.title = self._nombre_hoja

            columnas = RegistroPlano.columnas()
            hoja.append(list(columnas))
            self._dar_formato_encabezado(hoja, len(columnas))

            for registro in registros:
                hoja.append([self._valor_para_celda(valor) for valor in registro.valores()])

            self._ajustar_columnas(hoja, columnas, registros)
            hoja.freeze_panes = "A2"
            hoja.auto_filter.ref = (
                f"A1:{get_column_letter(len(columnas))}{max(len(registros) + 1, 1)}"
            )

            libro.save(destino)
        except ErrorDeExportacion:
            raise
        except OSError as error:
            raise ErrorDeExportacion(f"No fue posible escribir el archivo {destino}: {error}") from error
        except Exception as error:  # pragma: no cover - defensa ante fallos de openpyxl
            raise ErrorDeExportacion(f"Fallo al generar el Excel: {error}") from error

        logger.info("Excel escrito con %s filas de datos en %s.", len(registros), destino)
        return destino

    # -- Formato ---------------------------------------------------------------

    @staticmethod
    def _dar_formato_encabezado(hoja: Worksheet, total_columnas: int) -> None:
        relleno = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
        fuente = Font(bold=True, color="FFFFFF")
        alineacion = Alignment(horizontal="center", vertical="center")
        for indice in range(1, total_columnas + 1):
            celda = hoja.cell(row=1, column=indice)
            celda.fill = relleno
            celda.font = fuente
            celda.alignment = alineacion

    @staticmethod
    def _ajustar_columnas(
        hoja: Worksheet, columnas: Sequence[str], registros: Sequence[RegistroPlano]
    ) -> None:
        for indice, nombre in enumerate(columnas, start=1):
            ancho = len(nombre)
            for registro in registros:
                valor = registro.como_diccionario()[nombre]
                if valor is None:
                    continue
                primera_linea = str(valor).split("\n", 1)[0]
                ancho = max(ancho, len(primera_linea))
            hoja.column_dimensions[get_column_letter(indice)].width = min(
                max(ancho + 2, ANCHO_MINIMO), ANCHO_MAXIMO
            )

    @staticmethod
    def _valor_para_celda(valor: Any) -> Any:
        """Sanea el texto: Excel rechaza caracteres de control."""
        if isinstance(valor, str):
            return "".join(caracter for caracter in valor if caracter not in CARACTERES_ILEGALES)
        return valor
